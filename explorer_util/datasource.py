import requests
import numpy as np
from PIL import Image as pilimage
from io import BytesIO
import base64
from numpy import genfromtxt
from openpyxl import load_workbook


class DataSource(object):
    """
    Generic data reader/writer
    """

    def __init__(self, source=None, source_type=None):
        self.source = source

    def load(self, filename):
        data = genfromtxt(filename, delimiter=',')
        return data

    @staticmethod
    def read_from_file_source(file, nbytes=None):
        """
        Read from file. Optionally the first nbytes from the file.
        :param file: file to be read
        :param nbytes: number of bytes to be read
        :return: binary contents of the file
        """
        with open(file, "rb") as f:
            if nbytes is not None:
                data = f.read(nbytes)
            else:
                data = f.read()
        return data

    @staticmethod
    def fetch_from_online_source(url):
        """
        Read from online source.
        :param url: url
        :return: contents of online source
        """
        response = requests.get(url)
        return response.content


class DataDecoder(object):
    """
    Data decoder
    """

    def __init__(self, data):
        self.data = data

    def decode(self, callback=None):
        """
        Decode data stream. If callback is None, the delegator will determine the callback method.
        :param callback:
        :return:
        """
        runner = CallbackRunner(input=self.data)
        if callback is not None:
            runner.add(callback)
        return runner.add(decode_image_to_array).run()

    def inspect(self, buffer=None):
        """
        inspect buffer to determine data format
        :return:
        """
        raise NotImplementedError("inspect() API method not yet implemented")


def decode_image_to_array(buf: bytes):
    return _process_data_buffer(buf, _decode_image_to_array)


def base64_decode(data):
    return base64.b64decode(data)


class CallbackRunner(object):

    def __init__(self, input=None, callback=None):
        self.data = input
        self.callbacks = [callback] if callback is not None else []

    def add(self, cb):
        self.callbacks.append(cb)
        return self

    def run(self):
        if self.data is not None:
            processed_data = self.data
            for f in self.callbacks:
                processed_data = f(processed_data)
            return processed_data


def get_columns_from_xlsx_workbook(wb):
    allrows = np.array([row for row in wb.active.iter_rows(min_row=2, max_col=2, max_row=50, values_only=True)])
    return np.array([i for i in allrows if i[0] is not None])


def load_excel_doc_from_buf(buf: bytes):
    return _process_data_buffer(buf, _load_excel_doc_from_buf)


def _decode_image_to_array(bytesio: BytesIO):
    """
    Image decoder
    :param buf: memory buffer with image data
    :type buf:
    :return: array [width, height, cc] of image pixel data, with cc: colorchannels.
    :rtype: numpy.ndarray
    """
    image_object = pilimage.open(bytesio, mode="r")
    return np.array(image_object).astype(np.float32)


def _load_excel_doc_from_buf(bytesio: BytesIO):
    return load_workbook(bytesio)


def _process_data_buffer(buf: bytes, func):
    memory_stream = BytesIO(buf)
    try:
        memory_stream.seek(0)
        return func(memory_stream)
    except IOError as io_error:
        raise IOError("Unable to read buffer: {0}".format(io_error))
    except:
        raise IOError("Unexpected error while reading buffer")
    finally:
        if not memory_stream.closed:
            memory_stream.close()